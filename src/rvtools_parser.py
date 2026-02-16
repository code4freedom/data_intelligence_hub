import os
import logging
import json
import hashlib
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Generator, Iterable, List, Optional

import openpyxl
import pandas as pd

logger = logging.getLogger("rvtools_parser")
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")


def stream_sheet_rows(xlsx_path: str, sheet_name: str) -> Generator[Dict, None, None]:
    """Stream rows from an XLSX sheet as dictionaries (header -> value).

    Uses openpyxl in read_only mode so memory stays bounded for large files.
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)

    try:
        headers = next(rows)
    except StopIteration:
        return

    # normalize headers
    norm_headers = []
    for i, h in enumerate(headers):
        if h is None:
            norm_headers.append(f"col_{i}")
        else:
            norm_headers.append(str(h))

    for r in rows:
        yield {k: v for k, v in zip(norm_headers, r)}


def _md5_file(path: Path) -> str:
    h = hashlib.md5()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def _upload_file_s3(local_path: Path, bucket: str, key: str, endpoint_url: str, access_key: str, secret_key: str):
    import boto3
    s3 = boto3.resource("s3", endpoint_url=endpoint_url, aws_access_key_id=access_key, aws_secret_access_key=secret_key)
    # create bucket if not exists (MinIO tolerates create)
    try:
        s3.create_bucket(Bucket=bucket)
    except Exception:
        pass
    s3.Bucket(bucket).upload_file(str(local_path), key)


def chunk_and_write(xlsx_path: str, sheet_name: str, out_dir: str, chunk_size: int = 5000,
                    upload_s3: bool = False, s3_endpoint: Optional[str] = None,
                    s3_access_key: Optional[str] = None, s3_secret_key: Optional[str] = None,
                    s3_bucket: Optional[str] = None, ingest_id: Optional[str] = None) -> Dict:
    """Parse `sheet_name` from `xlsx_path` and write Parquet chunks to `out_dir`.

    If `upload_s3` is True the chunks are uploaded to the S3-compatible store and a manifest JSON
    describing the chunks is returned (and written locally and uploaded).
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    written = []
    buffer = []
    count = 0
    chunk_idx = 0
    # support optional shard-by-key mode
    shards = {}
    total_rows = 0

    for row in stream_sheet_rows(xlsx_path, sheet_name):
        buffer.append(row)
        count += 1
        total_rows += 1
        if count >= chunk_size:
            out_path = out_dir / f"chunk_{sheet_name}_{chunk_idx:06d}.parquet"
            df = pd.DataFrame(buffer)
            df.to_parquet(out_path, engine="pyarrow", index=False)
            rows = len(buffer)
            size = out_path.stat().st_size
            md5 = _md5_file(out_path)
            entry = {"name": out_path.name, "rows": rows, "size": size, "md5": md5, "local_path": str(out_path)}
            if upload_s3 and s3_bucket and s3_endpoint and s3_access_key and s3_secret_key:
                key = f"{ingest_id or 'ingest'}/chunks/{out_path.name}"
                _upload_file_s3(out_path, s3_bucket, key, s3_endpoint, s3_access_key, s3_secret_key)
                entry["s3_key"] = key
                entry["s3_bucket"] = s3_bucket
            written.append(entry)
            logger.info("Wrote chunk %s (%d rows)", out_path, rows)
            buffer = []
            count = 0
            chunk_idx += 1

    # final partial chunk
    if buffer:
        out_path = out_dir / f"chunk_{sheet_name}_{chunk_idx:06d}.parquet"
        df = pd.DataFrame(buffer)
        df.to_parquet(out_path, engine="pyarrow", index=False)
        rows = len(buffer)
        size = out_path.stat().st_size
        md5 = _md5_file(out_path)
        entry = {"name": out_path.name, "rows": rows, "size": size, "md5": md5, "local_path": str(out_path)}
        if upload_s3 and s3_bucket and s3_endpoint and s3_access_key and s3_secret_key:
            key = f"{ingest_id or 'ingest'}/chunks/{out_path.name}"
            _upload_file_s3(out_path, s3_bucket, key, s3_endpoint, s3_access_key, s3_secret_key)
            entry["s3_key"] = key
            entry["s3_bucket"] = s3_bucket
        written.append(entry)
        logger.info("Wrote final chunk %s (%d rows)", out_path, rows)


    manifest = {
        "ingest_id": ingest_id or "ingest",
        "sheet": sheet_name,
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "chunk_count": len(written),
        "total_rows": total_rows,
        "chunks": written,
    }

    manifest_name = f"manifest_{sheet_name}_{manifest['ingest_id']}.json"
    manifest_path = out_dir / manifest_name
    with manifest_path.open("w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2)

    # upload manifest
    if upload_s3 and s3_bucket and s3_endpoint and s3_access_key and s3_secret_key:
        key = f"{ingest_id or 'ingest'}/{manifest_name}"
        _upload_file_s3(manifest_path, s3_bucket, key, s3_endpoint, s3_access_key, s3_secret_key)
        manifest["s3_manifest_key"] = key
        manifest["s3_bucket"] = s3_bucket

    return {"manifest_path": str(manifest_path), "manifest": manifest}


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="RVTools sheet stream chunker with optional S3 upload and manifest")
    parser.add_argument("--input", required=True, help="path to RVTools XLSX file")
    parser.add_argument("--sheet", required=True, help="sheet name to parse")
    parser.add_argument("--out", required=True, help="output directory for chunks and manifest")
    parser.add_argument("--chunk-size", type=int, default=5000, help="rows per chunk")
    parser.add_argument("--upload-s3", action="store_true", help="upload chunks and manifest to S3/MinIO")
    parser.add_argument("--s3-endpoint", default=os.environ.get("S3_ENDPOINT"), help="S3 endpoint URL (MinIO)")
    parser.add_argument("--s3-access-key", default=os.environ.get("S3_ACCESS_KEY"), help="S3 access key")
    parser.add_argument("--s3-secret-key", default=os.environ.get("S3_SECRET_KEY"), help="S3 secret key")
    parser.add_argument("--s3-bucket", default=os.environ.get("S3_BUCKET"), help="S3 bucket")
    parser.add_argument("--ingest-id", default=os.environ.get("INGEST_ID"), help="ingest id / prefix")
    parser.add_argument("--shards", type=int, default=0, help="number of hash shards (0=disabled)")
    parser.add_argument("--shard-key", default=None, help="column name to use as shard key (e.g. Name or Host)")

    args = parser.parse_args()
    result = chunk_and_write(args.input, args.sheet, args.out, args.chunk_size,
                             upload_s3=args.upload_s3,
                             s3_endpoint=args.s3_endpoint,
                             s3_access_key=args.s3_access_key,
                             s3_secret_key=args.s3_secret_key,
                             s3_bucket=args.s3_bucket,
                             ingest_id=args.ingest_id)
    # print manifest path as JSON for callers
    print(json.dumps(result))
